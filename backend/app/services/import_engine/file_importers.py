from __future__ import annotations

import csv
import io
import json
from abc import ABC, abstractmethod
from collections.abc import Iterable, Iterator

from fastapi import UploadFile


class BaseFileImporter(ABC):
    @abstractmethod
    def rows(self, file: UploadFile) -> Iterable[dict[str, object]]:
        raise NotImplementedError


class CSVImporter(BaseFileImporter):
    def rows(self, file: UploadFile) -> Iterator[dict[str, object]]:
        file.file.seek(0)
        wrapped = io.TextIOWrapper(file.file, encoding="utf-8", newline="")
        reader = csv.DictReader(wrapped)
        for row in reader:
            yield {k: v for k, v in row.items()}


class JSONImporter(BaseFileImporter):
    def rows(self, file: UploadFile) -> Iterator[dict[str, object]]:
        file.file.seek(0)
        raw = file.file.read()
        payload = json.loads(raw)
        if isinstance(payload, list):
            for row in payload:
                if isinstance(row, dict):
                    yield row
        elif isinstance(payload, dict):
            records = payload.get("data", [])
            for row in records:
                if isinstance(row, dict):
                    yield row


class ExcelImporter(BaseFileImporter):
    def rows(self, file: UploadFile) -> Iterator[dict[str, object]]:
        try:
            from openpyxl import load_workbook
        except Exception as exc:  # pragma: no cover - dependency/environment sensitive
            raise RuntimeError("Excel import requires openpyxl") from exc

        file.file.seek(0)
        wb = load_workbook(file.file, read_only=True)
        ws = wb.active
        headers: list[str] = []
        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if idx == 1:
                headers = [str(col) if col is not None else "" for col in row]
                continue
            yield {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
