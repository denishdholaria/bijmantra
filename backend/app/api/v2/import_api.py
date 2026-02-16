"""Generic Data Import API (CSV/Excel/JSON)."""

from __future__ import annotations

import csv
import io
import json
from typing import Any

from fastapi import APIRouter, BackgroundTasks, Depends, File, Form, HTTPException, UploadFile
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import get_current_user, get_db
from app.core.database import AsyncSessionLocal
from app.models.core import User
from app.models.import_job import ImportJob
from app.schemas.imports import ImportJobResponse, ImportUploadResponse
from app.services.import_engine import (
    CSVImporter,
    ExcelImporter,
    GermplasmImporter,
    JSONImporter,
    ObservationImporter,
    TrialImporter,
)


router = APIRouter(prefix="/import", tags=["Data Import"])

SUPPORTED_MIME_TYPES = {
    "text/csv",
    "application/csv",
    "application/vnd.ms-excel",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/json",
    "text/json",
}


def _pick_file_importer(upload: UploadFile):
    filename = (upload.filename or "").lower()
    if filename.endswith(".csv"):
        return CSVImporter()
    if filename.endswith(".xlsx") or filename.endswith(".xls"):
        return ExcelImporter()
    if filename.endswith(".json"):
        return JSONImporter()

    if upload.content_type in {"text/csv", "application/csv"}:
        return CSVImporter()
    if upload.content_type in {
        "application/vnd.ms-excel",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }:
        return ExcelImporter()
    if upload.content_type in {"application/json", "text/json"}:
        return JSONImporter()

    raise HTTPException(status_code=400, detail=f"Unsupported upload type: {upload.content_type}")


def _pick_domain_importer(import_type: str, db: AsyncSession, org_id: int, user_id: int):
    normalized = import_type.strip().lower()
    if normalized == "germplasm":
        return GermplasmImporter(db, org_id, user_id)
    if normalized == "trial":
        return TrialImporter(db, org_id, user_id)
    if normalized in {"observation", "observations"}:
        return ObservationImporter(db, org_id, user_id)
    raise HTTPException(status_code=400, detail=f"Unsupported import_type '{import_type}'")


def _extract_headers(upload: UploadFile) -> list[str]:
    filename = (upload.filename or "").lower()
    upload.file.seek(0)
    if filename.endswith(".csv"):
        sample = upload.file.read(8192).decode("utf-8", errors="ignore")
        upload.file.seek(0)
        reader = csv.reader(io.StringIO(sample))
        return next(reader, [])
    if filename.endswith(".json"):
        payload = json.loads(upload.file.read())
        upload.file.seek(0)
        if isinstance(payload, list) and payload and isinstance(payload[0], dict):
            return list(payload[0].keys())
        if isinstance(payload, dict) and isinstance(payload.get("data"), list) and payload["data"]:
            return list(payload["data"][0].keys())
        return []

    try:
        from openpyxl import load_workbook

        wb = load_workbook(upload.file, read_only=True)
        ws = wb.active
        first = next(ws.iter_rows(values_only=True), [])
        upload.file.seek(0)
        return [str(x) if x is not None else "" for x in first]
    except Exception:
        upload.file.seek(0)
        return []


async def _run_job(
    job_id: int,
    import_type: str,
    mapping: dict[str, str],
    formulas: dict[str, str],
    dry_run: bool,
    file_name: str,
    content_type: str,
    content_bytes: bytes,
    organization_id: int,
    user_id: int,
):
    async with AsyncSessionLocal() as db:
        try:
            job = await db.get(ImportJob, job_id)
            if not job:
                return

            job.status = "running"
            await db.commit()

            temp = UploadFile(filename=file_name, file=io.BytesIO(content_bytes))
            importer = _pick_file_importer(temp)
            domain_importer = _pick_domain_importer(import_type, db, organization_id, user_id)

            rows_iter = importer.rows(temp)
            report = await domain_importer.import_data(rows_iter, mapping=mapping, formulas=formulas, dry_run=dry_run)
            job.total_rows = report.success_count + len(report.errors)

            job.status = "completed" if not report.errors else "failed"
            job.success_count = report.success_count
            job.error_count = len(report.errors)
            job.report = report.model_dump()
            await db.commit()
        except Exception as exc:
            job = await db.get(ImportJob, job_id)
            if job:
                job.status = "failed"
                job.error_details = str(exc)
                await db.commit()


@router.post("/upload", response_model=ImportUploadResponse)
async def upload_import(
    background_tasks: BackgroundTasks,
    import_type: str = Form(...),
    mapping_config: str | None = Form(default=None),
    formulas: str | None = Form(default=None),
    dry_run: bool = Form(default=False),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if file.content_type not in SUPPORTED_MIME_TYPES:
        raise HTTPException(status_code=400, detail=f"Unsupported MIME type: {file.content_type}")

    mapping = json.loads(mapping_config) if mapping_config else {}
    formula_map = json.loads(formulas) if formulas else {}

    domain_importer = _pick_domain_importer(import_type, db, current_user.organization_id, current_user.id)
    headers = _extract_headers(file)
    suggestions = await domain_importer.suggest_mappings(headers)

    job = ImportJob(
        organization_id=current_user.organization_id,
        user_id=current_user.id,
        import_type=import_type.lower(),
        file_name=file.filename or "upload.dat",
        status="pending",
        dry_run=dry_run,
        mapping_config=mapping,
    )
    db.add(job)
    await db.flush()
    job_id = job.id
    await db.commit()

    file.file.seek(0)
    content_bytes = file.file.read()

    background_tasks.add_task(
        _run_job,
        job_id,
        import_type,
        mapping,
        formula_map,
        dry_run,
        file.filename or "upload.dat",
        file.content_type or "application/octet-stream",
        content_bytes,
        current_user.organization_id,
        current_user.id,
    )

    return ImportUploadResponse(job_id=job_id, status="queued", mapping_suggestions=suggestions)


@router.get("/jobs", response_model=list[ImportJobResponse])
async def list_import_jobs(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    result = await db.execute(
        select(ImportJob).where(ImportJob.organization_id == current_user.organization_id).order_by(ImportJob.id.desc())
    )
    jobs = result.scalars().all()
    return [
        ImportJobResponse(
            id=j.id,
            import_type=j.import_type,
            file_name=j.file_name,
            status=j.status,
            dry_run=j.dry_run,
            total_rows=j.total_rows or 0,
            success_count=j.success_count or 0,
            error_count=j.error_count or 0,
        )
        for j in jobs
    ]


@router.get("/template/{import_type}")
async def download_template(import_type: str) -> dict[str, Any]:
    templates = {
        "germplasm": ["germplasm_name", "accession_number", "genus", "species", "common_crop_name"],
        "trial": ["trial_name", "program_name", "location_name", "start_date", "end_date"],
        "observation": ["observation_unit_name", "trait", "value", "date"],
    }
    key = import_type.lower()
    if key not in templates:
        raise HTTPException(status_code=404, detail="Unknown import template")
    return {"import_type": key, "headers": templates[key]}
